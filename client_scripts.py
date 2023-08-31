import numpy as np
import pandas as pd
from openpyxl import load_workbook
import json
from os import path
import re
import glob


def extractNumber(filePath):
    pattern = r'\d+'

    numbers = re.findall(pattern, filePath)
    if numbers:
        return [int(numbers[0]), int(numbers[1]),int(numbers[2])]
    else:
        return None
    
def read_file(filename,duration,rps):
    try:
        json_data = {}
        with open(filename) as data:
            json_data = json.load(data)

        total = int(duration) * int(rps) * int(json_data['options']['connections'])
        data = {
            'total_requests':total,
            'processed_log_file': filename,
            'clients': json_data['options']['connections'],
            'workers': json_data['options']['concurrency'],
            'request_count': json_data['options']['total'],
            'min_time': f"{round(json_data['fastest'] / 1000000, 4)}ms",
            'max_time': f"{round(json_data['slowest']  / 1000000, 4)}ms",
            'mean': f"{round(json_data['average']  / 1000000, 4)}ms",
            'rps': f"{round(json_data['options']['rps'], 2)}",
            "success_request": json_data['count'],
            "duration": f"{round(json_data['options']['duration'] / 1000000000,3)}s"
        }

        for percentage in json_data['latencyDistribution']:
            print(percentage)
            if percentage['percentage'] == 50:
                data["50th_percentile"] = f"{round(percentage['latency'] / 1000000, 4)}ms"
            if percentage['percentage'] == 90:
                data["90th_percentile"] = f"{round(percentage['latency'] / 1000000, 4)}ms"
            if percentage['percentage'] == 99:
                data["99th_percentile"] = f"{round(percentage['latency'] / 1000000, 4)}ms"

        return data
    except FileNotFoundError:
        print(f"File Not Found {filename}")
    except Exception as e:
        print(f"An error occurred while reading json file: {e}")
    return {}


def addToJSONFile(file_path, latencies_data):
    try:
        print("file add to json file")
        if path.isfile(file_path) is False:
            print("file opened")
            fp = open(file_path, "w")
            fp.close()
            print("file closed")

        dict_obj = []
        with open(file_path) as fp:
            try:
                dict_obj = json.load(fp)
            except json.JSONDecodeError:
                pass

        dict_obj.append(latencies_data)

        with open(file_path, 'w') as json_file:
            json.dump(dict_obj, json_file,
                      indent=2,
                      separators=(',', ': '))

        print('Successfully written to the JSON file\n\n')
    except Exception as e:
        print(f"An error occurred addToJSONFile: {e}")


def convertJSONToArray(item):
    return [
        item['total_requests'],
        item['success_request'],
        item['clients'],
        item['duration'],
        item['rps'],
        item['min_time'],
        item['mean'],
        item['max_time'],
        item["50th_percentile"],
        item["90th_percentile"],
        item["99th_percentile"],
    ]


def addJSONToExcel(inputJSONFile, outputXlsxFile):
    try:
        json_file = open(inputJSONFile)
        data = json.load(json_file)
        wb = load_workbook(outputXlsxFile)
        sheet_name = 'Sheet3'  # Change this to the desired sheet name
        sheet = wb[sheet_name]
        index = sheet.max_row + 1
        for item in data:
            try:
                row = convertJSONToArray(item)
                row.insert(0, index)
                sheet.append(row)
                index += 1
            except Exception as e:
                print(f"Error while appending row to excel {item} {e}")
        wb.save(outputXlsxFile)
    except Exception as e:
        print(f"An error occurred addJSONToExcel: {e}")


def extractNumber(filePath):
    pattern = r'\d+'

    numbers = re.findall(pattern, filePath)
    if numbers:
        return [int(numbers[0]), int(numbers[1]),int(numbers[2])]
    else:
        return None


def main():
    try:
        # Find all files matching the pattern "output_*.log"
        file_pattern = "./output/logs/output_client_*.json"
        matching_files = glob.glob(file_pattern)
        output_file = './output/json/result_client.json'

        if not matching_files:
            print('No Matching files found')
            return
        # Read the content of each matching file
        for file_path in matching_files:
            print("file = ", file_path)
            [clients,duration, rps] = extractNumber(filePath=file_path)
            result = read_file(filename=file_path,duration=duration,rps=rps)
            print("result = ", result)
            if result:
                addToJSONFile(file_path=output_file, latencies_data=result)
            else:
                print(f"File Empty {file_path}")

        print(
            f'** Processed logs file & result added to JSON File {output_file}**')
        excel_file = 'load_test_grpc.xlsx'
        addJSONToExcel(inputJSONFile=output_file, outputXlsxFile=excel_file)
        print(f'Succesfully added to Excel file {excel_file}')
    except Exception as e:
        print(f"An error occurred main method : {e}")


main()


# convert output json from ghz commands into extract single json file for all result
# then add excel sheet
