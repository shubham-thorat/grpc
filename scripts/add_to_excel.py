import numpy as np
import pandas as pd
from openpyxl import load_workbook
import json
from os import path
import re
import glob


def extractNumber(filePath):
    pattern = r'\d+'

    numbers = re.findall(pattern, filePath)
    if numbers:
        return [int(numbers[0]), int(numbers[1]),int(numbers[2])]
    else:
        return None
    
def read_file(filename):
    try:
        json_data = {}
        with open(filename) as data:
            json_data = json.load(data)

        data = {
            "connections": json_data['options']['connections'],
            "concurrency": json_data['options']['concurrency'],
            "duration": f"{round(json_data['options']['duration'] * 1e-6,2)}",
            "req_per_sec_input":f"{round(json_data['options']['rps'],3)}",
            "req_per_sec":f"{round(json_data['rps'],3)}",
            "min_latency":f"{round(json_data['fastest'] * 1e-6,3)}",
            "max_latency":f"{round(json_data['slowest'] * 1e-6,3)}",
        }

        for percentage in json_data['latencyDistribution']:
            print(percentage)
            if percentage['percentage'] == 50:
                data["50th_percentile"] = f"{round(percentage['latency'] * 1e-6, 3)}"
            if percentage['percentage'] == 90:
                data["90th_percentile"] = f"{round(percentage['latency'] * 1e-6, 3)}"
            if percentage['percentage'] == 99:
                data["99th_percentile"] = f"{round(percentage['latency'] * 1e-6, 3)}"

        return data
    except FileNotFoundError:
        print(f"File Not Found {filename}")
    except Exception as e:
        print(f"An error occurred while reading json file: {e}")
    return {}


def addToJSONFile(file_path, latencies_data):
    try:
        
        if path.isfile(file_path) is False:
            fp = open(file_path, "w")
            fp.close()
        
        dict_obj = []
        with open(file_path) as fp:
            try:
                dict_obj = json.load(fp)
            except json.JSONDecodeError:
                pass

        dict_obj.append(latencies_data)

        with open(file_path, 'w') as json_file:
            json.dump(dict_obj, json_file,
                      indent=2,
                      separators=(',', ': '))

        print('Successfully written to the JSON file\n\n')
    except Exception as e:
        print(f"An error occurred addToJSONFile: {e}")


def convertJSONToArray(index,item):
    return [
        index,
        item['connections'],
        item['concurrency'],
        item['duration'],
        item['req_per_sec_input'],
        item['req_per_sec'],
        item['50th_percentile'],
        item['90th_percentile'],
        item['99th_percentile'],
        item["min_latency"],
        item["max_latency"]
    ]


def addJSONToExcel(inputJSONFile, outputXlsxFile):
    try:
        json_file = open(inputJSONFile)
        data = json.load(json_file)
        wb = load_workbook(outputXlsxFile)
        sheet_name = 'GRPC'  # Change this to the desired sheet name
        sheet = wb[sheet_name]
        index = sheet.max_row + 1
        sheet.append([])
        print("Starting row : ",index)
        index += 1
        for item in data:
            try:
                row = convertJSONToArray(index,item)
                sheet.append(row)
                index += 1
            except Exception as e:
                print(f"Error while appending row to excel {item} {e}")
        wb.save(outputXlsxFile)
    except Exception as e:
        print(f"An error occurred addJSONToExcel: {e}")


def extractNumber(filePath):
    pattern = r'\d+'

    numbers = re.findall(pattern, filePath)
    if numbers:
        return [int(numbers[0]), int(numbers[1]),int(numbers[2])]
    else:
        return None


def main():
    try:
        # Find all files matching the pattern "output_*.log"
        file_pattern = "./result/output_client_*.json"
        matching_files = glob.glob(file_pattern)

        if not matching_files:
            print('No Matching files found')
            return
        
        output_file = './output.json'
        # Read the content of each matching file
        for file_path in matching_files:
            result = read_file(filename=file_path)
            if result:
                addToJSONFile(file_path=output_file, latencies_data=result)
            else:
                print(f"File Empty {file_path}")

        excel_file = 'Load_Test.xlsx'
        addJSONToExcel(inputJSONFile=output_file, outputXlsxFile=excel_file)
        print(f'Succesfully added to Excel file {excel_file}')
    except Exception as e:
        print(f"An error occurred main method : {e}")


main()


# convert output json from ghz commands into extract single json file for all result
# then add excel sheet
